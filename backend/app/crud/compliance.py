# app/services/compliance_service.py
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import datetime, date, timedelta
from typing import Dict, Any, List
from ..models import Employee, Department, Training, Enrollment, Certification
from ..schemas import ComplianceMetrics, DepartmentCompliance, CertificationStatus, UpcomingExpiration, MissingCertification
import pandas as pd
from io import BytesIO

class CRUDCompliance:
    def __init__(self):
        pass
    
    def _to_date(self, dt_or_date):
        """Convert datetime or date to date object"""
        if isinstance(dt_or_date, datetime):
            return dt_or_date.date()
        return dt_or_date
    
    def get_compliance_report(self, db: Session, filters: Dict[str, Any]) -> ComplianceMetrics:
        """Generate comprehensive compliance report"""
        
        # Get all employees (apply department filter)
        query = db.query(Employee)
        
        if filters.get('department') and filters['department'] != 'all':
            query = query.join(Department).filter(Department.name == filters['department'])
        
        employees = query.all()
        total_employees = len(employees)
        
        # Calculate compliance metrics
        compliance_data = self._calculate_compliance_metrics(db, employees)
        
        # Get department-wise compliance
        department_compliance = self._get_department_compliance(db, filters)
        
        # Get certification status
        certification_status = self._get_certification_status(db, filters)
        
        # Get upcoming expirations
        upcoming_expirations = self._get_upcoming_expirations(db, filters)
        
        # Get missing certifications
        missing_certifications = self._get_missing_certifications(db, filters)
        
        # Get training statistics
        training_stats = self._get_training_statistics(db, filters)
        
        return ComplianceMetrics(
            total_employees=total_employees,
            compliant_employees=compliance_data['compliant_employees'],
            non_compliant_employees=compliance_data['non_compliant_employees'],
            expiring_soon=compliance_data['expiring_soon'],
            expired_certifications=compliance_data['expired_certifications'],
            total_trainings=training_stats['total_trainings'],
            completed_trainings=training_stats['completed_trainings'],
            pending_trainings=training_stats['pending_trainings'],
            overall_compliance_rate=compliance_data['overall_compliance_rate'],
            department_compliance=department_compliance,
            certification_status=certification_status,
            upcoming_expirations=upcoming_expirations,
            missing_certifications=missing_certifications
        )
    
    def _calculate_compliance_metrics(self, db: Session, employees: List[Employee]) -> Dict[str, Any]:
        """Calculate overall compliance metrics"""
        
        today = datetime.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        compliant_employees = 0
        expiring_soon = 0
        expired_certifications = 0
        
        for employee in employees:
            # Get employee certifications
            certs = db.query(Certification).filter(
                Certification.employee_id == employee.id
            ).all()
            
            # Check if employee has valid certifications
            has_valid_certs = all(
                cert.status == "active" and 
                (not cert.expires_at or self._to_date(cert.expires_at) >= today)
                for cert in certs
            ) if certs else False
            
            # Check for expiring/expired certs
            has_expiring = any(
                cert.expires_at and 
                today <= self._to_date(cert.expires_at) <= thirty_days_from_now 
                for cert in certs
            )

            has_expired = any(
                cert.expires_at and 
                self._to_date(cert.expires_at) < today 
                for cert in certs
            )
            
            # Check training completion
            enrollments = db.query(Enrollment).filter(
                Enrollment.employee_id == employee.id
            ).all()
            
            has_completed_trainings = all(
                enrollment.status == "completed" 
                for enrollment in enrollments
            ) if enrollments else False
            
            # Employee is compliant if they have valid certs and completed trainings
            if has_valid_certs and has_completed_trainings:
                compliant_employees += 1
            
            if has_expiring:
                expiring_soon += 1
            
            if has_expired:
                expired_certifications += 1
        
        non_compliant_employees = len(employees) - compliant_employees
        overall_compliance_rate = (compliant_employees / len(employees)) * 100 if employees else 0
        
        return {
            'compliant_employees': compliant_employees,
            'non_compliant_employees': non_compliant_employees,
            'expiring_soon': expiring_soon,
            'expired_certifications': expired_certifications,
            'overall_compliance_rate': round(overall_compliance_rate, 2)
        }
    
    def _get_department_compliance(self, db: Session, filters: Dict[str, Any]) -> List[DepartmentCompliance]:
        """Get compliance statistics by department"""
        
        departments = db.query(Department).all()
        department_compliance = []
        
        for dept in departments:
            # Apply department filter
            if filters.get('department') and filters['department'] != 'all':
                if dept.name != filters['department']:
                    continue
            
            employees = db.query(Employee).filter(
                Employee.department_id == dept.id
            ).all()
            
            if not employees:
                continue
            
            compliant_count = 0
            completed_trainings_count = 0
            pending_trainings_count = 0
            total_trainings_count = 0
            
            for emp in employees:
                # Get certifications
                certs = db.query(Certification).filter(
                    Certification.employee_id == emp.id
                ).all()
                
                # Get enrollments
                enrollments = db.query(Enrollment).filter(
                    Enrollment.employee_id == emp.id
                ).all()
                
                # Check valid certifications
                has_valid_certs = all(
                    cert.status == "active" and 
                    (not cert.expires_at or self._to_date(cert.expires_at) >= datetime.now().date())
                    for cert in certs
                ) if certs else False
                
                # Check completed trainings
                has_completed_trainings = all(
                    enrollment.status == "completed" 
                    for enrollment in enrollments
                ) if enrollments else False
                
                # Count training statistics
                if enrollments:
                    completed_trainings_count += sum(1 for e in enrollments if e.status == "completed")
                    pending_trainings_count += sum(1 for e in enrollments if e.status in ["enrolled", "in_progress"])
                    total_trainings_count += len(enrollments)
                
                if has_valid_certs and has_completed_trainings:
                    compliant_count += 1

            compliance_rate = (compliant_count / len(employees)) * 100 if employees else 0
            
            department_compliance.append(
                DepartmentCompliance(
                    department=dept.name,
                    compliance_rate=round(compliance_rate, 2),
                    total_employees=len(employees),
                    compliant_employees=compliant_count,
                    completed_trainings=completed_trainings_count,
                    pending_trainings=pending_trainings_count,
                    total_trainings=total_trainings_count
                )
            )
        
        return department_compliance
    
    def _get_certification_status(self, db: Session, filters: Dict[str, Any]) -> List[CertificationStatus]:
        """Get certification status statistics"""
        
        today = datetime.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        # Build query
        query = (
            db.query(Certification)
            .join(Training, Certification.training_id == Training.id)
            .join(Employee, Certification.employee_id == Employee.id)
            .join(Department, Employee.department_id == Department.id)
        )
        
        # Apply department filter
        if filters.get('department') and filters['department'] != 'all':
            query = query.filter(Department.name == filters['department'])
        
        certifications = query.all()
        
        # Group by training name
        cert_by_training = {}
        for cert in certifications:
            training_name = cert.training.name if cert.training else "Unknown Training"
            
            if training_name not in cert_by_training:
                cert_by_training[training_name] = []
            cert_by_training[training_name].append(cert)

        certification_status = []
        
        for training_name, certs in cert_by_training.items():
            total = len(certs)
            
            # Count valid certifications
            valid = sum(1 for cert in certs 
                       if cert.status == "active" and 
                       (not cert.expires_at or self._to_date(cert.expires_at) >= today))
            
            # Count expiring soon
            expiring_soon = sum(1 for cert in certs 
                              if cert.expires_at and 
                              today <= self._to_date(cert.expires_at) <= thirty_days_from_now)
            
            # Count expired
            expired = sum(1 for cert in certs 
                         if cert.expires_at and 
                         self._to_date(cert.expires_at) < today)
            
            compliance_rate = (valid / total) * 100 if total > 0 else 0
            
            certification_status.append(
                CertificationStatus(
                    certification=training_name,
                    total=total,
                    valid=valid,
                    expiring_soon=expiring_soon,
                    expired=expired,
                    compliance_rate=round(compliance_rate, 2)
                )
            )
        
        return certification_status
    
    def _get_upcoming_expirations(self, db: Session, filters: Dict[str, Any]) -> List[UpcomingExpiration]:
        """Get certifications expiring soon"""
        
        today = datetime.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        # Build query
        query = (
            db.query(Certification)
            .join(Training, Certification.training_id == Training.id)
            .join(Employee, Certification.employee_id == Employee.id)
            .join(Department, Employee.department_id == Department.id)
            .filter(
                and_(
                    Certification.expires_at >= today,
                    Certification.expires_at <= thirty_days_from_now
                )
            )
        )
        
        # Apply department filter
        if filters.get('department') and filters['department'] != 'all':
            query = query.filter(Department.name == filters['department'])
        
        expiring_certs = query.all()
        
        upcoming_expirations = []
        for cert in expiring_certs:
            expires_at_date = self._to_date(cert.expires_at)
            days_until_expiry = (expires_at_date - today).days
            
            training_name = cert.training.name if cert.training else "Unknown Training"
            
            upcoming_expirations.append(
                UpcomingExpiration(
                    id=cert.id,
                    employee_name=f"{cert.employee.first_name} {cert.employee.last_name}",
                    certification_name=training_name,
                    expiry_date=expires_at_date,
                    days_until_expiry=days_until_expiry,
                    department=cert.employee.department.name if cert.employee.department else "N/A"
                )
            )
        
        return sorted(upcoming_expirations, key=lambda x: x.days_until_expiry)
    
    def _get_missing_certifications(self, db: Session, filters: Dict[str, Any]) -> List[MissingCertification]:
        """Get employees missing required certifications"""
        
        missing_certifications = []
        today = datetime.now().date()
        
        # Get all completed enrollments
        query = (
            db.query(Enrollment)
            .join(Employee, Enrollment.employee_id == Employee.id)
            .join(Training, Enrollment.training_id == Training.id)
            .join(Department, Employee.department_id == Department.id)
            .filter(Enrollment.status == "completed")
        )
        
        # Apply department filter
        if filters.get('department') and filters['department'] != 'all':
            query = query.filter(Department.name == filters['department'])
        
        completed_enrollments = query.all()
        
        for enrollment in completed_enrollments:
            # Check if employee has certification for this training
            has_cert = (
                db.query(Certification)
                .filter(
                    and_(
                        Certification.employee_id == enrollment.employee_id,
                        Certification.training_id == enrollment.training_id
                    )
                )
                .first()
            )
            
            if not has_cert:
                days_overdue = 0
                if enrollment.completed_date:
                    completion_date = self._to_date(enrollment.completed_date)
                    days_overdue = max(0, (today - completion_date).days - 30)
                
                missing_certifications.append(
                    MissingCertification(
                        id=enrollment.employee_id,
                        employee_name=f"{enrollment.employee.first_name} {enrollment.employee.last_name}",
                        required_certification=enrollment.training.name,
                        department=enrollment.employee.department.name if enrollment.employee.department else "N/A",
                        days_overdue=days_overdue
                    )
                )
        
        return missing_certifications
    
    def _get_training_statistics(self, db: Session, filters: Dict[str, Any]) -> Dict[str, int]:
        """Get training completion statistics"""
        
        query = db.query(Training)
        
        # Note: Training statistics should be calculated based on employees in filtered department
        total_trainings = query.count()
        
        # Get enrollments with department filter
        enrollments_query = db.query(Enrollment)
        
        if filters.get('department') and filters['department'] != 'all':
            enrollments_query = (
                enrollments_query
                .join(Employee, Enrollment.employee_id == Employee.id)
                .join(Department, Employee.department_id == Department.id)
                .filter(Department.name == filters['department'])
            )
        
        enrollments = enrollments_query.all()
        
        completed_trainings = sum(1 for e in enrollments if e.status == "completed")
        pending_trainings = sum(1 for e in enrollments if e.status in ["enrolled", "in_progress"])
        
        return {
            'total_trainings': total_trainings,
            'completed_trainings': completed_trainings,
            'pending_trainings': pending_trainings
        }
    
    def export_to_excel(self, db: Session, filters: Dict[str, Any]) -> BytesIO:
        """Export compliance report to Excel (.xlsx format)"""
        try:
            report = self.get_compliance_report(db, filters)
            output = BytesIO()
            
            # Create Excel writer WITHOUT problematic engine_kwargs
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Summary sheet
                summary_df = pd.DataFrame({
                    'Metric': [
                        'Total Employees', 'Compliant Employees', 'Non-Compliant Employees',
                        'Overall Compliance Rate', 'Expiring Soon', 'Expired Certifications',
                        'Total Trainings', 'Completed Trainings', 'Pending Trainings'
                    ],
                    'Value': [
                        report.total_employees, 
                        report.compliant_employees, 
                        report.non_compliant_employees,
                        f"{report.overall_compliance_rate}%", 
                        report.expiring_soon, 
                        report.expired_certifications,
                        report.total_trainings, 
                        report.completed_trainings, 
                        report.pending_trainings
                    ]
                })
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Department Compliance
                if hasattr(report, 'department_compliance') and report.department_compliance:
                    dept_data = []
                    for dept in report.department_compliance:
                        dept_data.append({
                            'Department': dept.department,
                            'Compliance Rate': f"{dept.compliance_rate}%",
                            'Total Employees': dept.total_employees,
                            'Compliant Employees': dept.compliant_employees,
                            'Completed Trainings': getattr(dept, 'completed_trainings', 0),
                            'Pending Trainings': getattr(dept, 'pending_trainings', 0),
                            'Total Trainings': getattr(dept, 'total_trainings', 0)
                        })
                    pd.DataFrame(dept_data).to_excel(writer, sheet_name='Department Compliance', index=False)
                
                # Certification Status
                if hasattr(report, 'certification_status') and report.certification_status:
                    cert_data = []
                    for cert in report.certification_status:
                        cert_data.append({
                            'Certification': cert.certification,
                            'Total': cert.total,
                            'Valid': cert.valid,
                            'Expiring Soon': cert.expiring_soon,
                            'Expired': cert.expired,
                            'Compliance Rate': f"{cert.compliance_rate}%"
                        })
                    pd.DataFrame(cert_data).to_excel(writer, sheet_name='Certification Status', index=False)
                
                # Upcoming Expirations
                if hasattr(report, 'upcoming_expirations') and report.upcoming_expirations:
                    expiry_data = []
                    for exp in report.upcoming_expirations:
                        expiry_data.append({
                            'Employee': exp.employee_name,
                            'Certification': exp.certification_name,
                            'Expiry Date': exp.expiry_date.strftime('%Y-%m-%d') if exp.expiry_date else "N/A",
                            'Days Until Expiry': exp.days_until_expiry,
                            'Department': exp.department
                        })
                    pd.DataFrame(expiry_data).to_excel(writer, sheet_name='Upcoming Expirations', index=False)
                
                # Missing Certifications
                if hasattr(report, 'missing_certifications') and report.missing_certifications:
                    missing_data = []
                    for missing in report.missing_certifications:
                        missing_data.append({
                            'Employee': missing.employee_name,
                            'Required Certification': missing.required_certification,
                            'Department': missing.department,
                            'Days Overdue': missing.days_overdue
                        })
                    pd.DataFrame(missing_data).to_excel(writer, sheet_name='Missing Certifications', index=False)
                
                # Auto-adjust column widths (optional)
                try:
                    for sheet_name in writer.sheets:
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                except Exception as e:
                    print(f"Warning: Could not adjust column widths: {e}")
            
            # CRITICAL: Seek to beginning before returning
            output.seek(0)
            
            # Verify we have data
            file_size = output.getbuffer().nbytes
            print(f"Created Excel file with size: {file_size} bytes")
            
            # Read first few bytes to check signature
            position = output.tell()
            first_bytes = output.read(4)
            output.seek(position)  # Reset position
            
            if first_bytes.startswith(b'PK'):
                print("✓ File has valid Excel (ZIP) signature")
            else:
                print(f"⚠ File signature: {first_bytes.hex()}")
            
            return output
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            import traceback
            traceback.print_exc()
            
            # Create a simple valid Excel file on error using openpyxl directly
            try:
                from openpyxl import Workbook
                
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Error"
                ws['A1'] = "Error occurred during export"
                ws['A2'] = str(e)
                wb.save(output)
                output.seek(0)
                
                print("Created error Excel file as fallback")
                return output
            except Exception as fallback_error:
                print(f"Even fallback failed: {fallback_error}")
                # Return empty bytes as last resort
                return BytesIO()
            
    def export_to_pdf(self, db: Session, filters: Dict[str, Any]) -> BytesIO:
        """Export compliance report to PDF using reportlab"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Get compliance report data
            report = self.get_compliance_report(db, filters)
            
            # Create PDF in memory
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                spaceBefore=20
            )
            
            # Story to hold all elements
            story = []
            
            # Title
            story.append(Paragraph("Compliance Report", title_style))
            
            # Report Period
            story.append(Paragraph(
                f"Report Period: {filters.get('date_range', {}).get('start', 'N/A')} to {filters.get('date_range', {}).get('end', 'N/A')}",
                styles['Normal']
            ))
            
            # Department Filter
            if filters.get('department') and filters['department'] != 'all':
                story.append(Paragraph(
                    f"Department: {filters['department']}",
                    styles['Normal']
                ))
            
            story.append(Spacer(1, 20))
            
            # Summary Section
            story.append(Paragraph("Summary", heading_style))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Employees', str(report.total_employees)],
                ['Compliant Employees', str(report.compliant_employees)],
                ['Non-Compliant Employees', str(report.non_compliant_employees)],
                ['Overall Compliance Rate', f"{report.overall_compliance_rate}%"],
                ['Expiring Soon', str(report.expiring_soon)],
                ['Expired Certifications', str(report.expired_certifications)],
                ['Total Trainings', str(report.total_trainings)],
                ['Completed Trainings', str(report.completed_trainings)],
                ['Pending Trainings', str(report.pending_trainings)],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(summary_table)
            
            # Department Compliance Section
            if report.department_compliance:
                story.append(Spacer(1, 20))
                story.append(Paragraph("Department Compliance", heading_style))
                
                dept_data = [['Department', 'Compliance Rate', 'Employees', 'Compliant']]
                for dept in report.department_compliance:
                    dept_data.append([
                        dept.department,
                        f"{dept.compliance_rate}%",
                        str(dept.total_employees),
                        str(dept.compliant_employees)
                    ])
                
                dept_table = Table(dept_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch])
                dept_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(dept_table)
            
            # Certification Status Section
            if report.certification_status:
                story.append(Spacer(1, 20))
                story.append(Paragraph("Certification Status", heading_style))
                
                cert_data = [['Certification', 'Total', 'Valid', 'Expiring Soon', 'Expired', 'Rate']]
                for cert in report.certification_status:
                    cert_data.append([
                        cert.certification[:30],  # Limit name length
                        str(cert.total),
                        str(cert.valid),
                        str(cert.expiring_soon),
                        str(cert.expired),
                        f"{cert.compliance_rate}%"
                    ])
                
                cert_table = Table(cert_data, colWidths=[2*inch, 0.8*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch])
                cert_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                story.append(cert_table)
            
            # Upcoming Expirations Section
            if report.upcoming_expirations:
                story.append(Spacer(1, 20))
                story.append(Paragraph("Upcoming Expirations (Next 30 Days)", heading_style))
                
                exp_data = [['Employee', 'Certification', 'Days Left', 'Department']]
                for exp in report.upcoming_expirations[:10]:  # Limit to 10
                    exp_data.append([
                        exp.employee_name,
                        exp.certification_name[:30],  # Limit name length
                        str(exp.days_until_expiry),
                        exp.department
                    ])
                
                exp_table = Table(exp_data, colWidths=[2*inch, 2*inch, 1*inch, 1.5*inch])
                exp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                story.append(exp_table)
            
            # Missing Certifications Section
            if report.missing_certifications:
                story.append(Spacer(1, 20))
                story.append(Paragraph("Missing Required Certifications", heading_style))
                
                miss_data = [['Employee', 'Required Certification', 'Department', 'Status']]
                for miss in report.missing_certifications[:10]:  # Limit to 10
                    status = f"{miss.days_overdue} days overdue" if miss.days_overdue > 0 else "Missing"
                    miss_data.append([
                        miss.employee_name,
                        miss.required_certification[:30],  # Limit name length
                        miss.department,
                        status
                    ])
                
                miss_table = Table(miss_data, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
                miss_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                story.append(miss_table)
            
            # Footer
            story.append(Spacer(1, 20))
            story.append(Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Italic']
            ))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer
            
        except ImportError:
            # If reportlab is not installed, fall back to Excel
            print("ReportLab not installed. Falling back to Excel export.")
            return self.export_to_excel(db, filters)
        except Exception as e:
            print(f"Error generating PDF: {e}")
            return self.export_to_excel(db, filters)

# Create instance
compliance = CRUDCompliance()